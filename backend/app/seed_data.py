from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = {
    "software": [
        "id", "name", "vendor", "category", "description", "capabilities",
        "industries", "deployment", "compliance", "license_model", "currency",
        "unit_license_cost", "maintenance_pct", "available_licenses",
        "assigned_licenses", "renewal_date", "status", "revenue_currency",
        "lifetime_revenue", "current_year_revenue", "pipeline_value",
        "closed_won_deals", "open_opportunities", "crm_product_id",
        "crm_last_sync",
    ],
    "customers": ["id", "name", "industry", "country", "currency", "contact_name", "contact_email"],
    "queries": [
        "id", "customer_id", "subject", "email_body", "received_at", "priority",
        "status", "budget", "currency", "preferred_deployment", "requirements",
    ],
    "users": ["id", "name", "email", "password", "role", "customer_id", "active"],
    "analyses": [
        "query_id", "software_id", "software_name", "score", "capability_score", "industry_score",
        "deployment_score", "budget_score", "compliance_score", "reasons", "gaps",
        "annual_cost", "currency", "summary", "extracted_requirements", "confidence",
        "source", "analyzed_at",
    ],
    "opportunities": [
        "id", "query_id", "customer_id", "software_id", "name", "owner",
        "stage", "probability", "amount", "currency", "expected_close",
        "source", "crm_id", "updated_at",
    ],
    "feedback": [
        "id", "query_id", "software_id", "user_id", "rating", "comment", "created_at",
    ],
    "alerts": [
        "id", "type", "severity", "title", "message", "entity_id", "due_date",
        "status", "created_at",
    ],
    "audit_logs": [
        "id", "user_id", "action", "entity_type", "entity_id", "details", "created_at",
    ],
    "settings": ["key", "value", "description"],
    "email_inbox": [
        "id", "provider", "external_id", "sender", "subject", "body",
        "received_at", "status", "query_id",
    ],
    "procurement": [
        "software_id", "annual_demand", "order_cost", "holding_cost_per_unit",
        "lead_time_days", "safety_stock", "reorder_point", "last_sale_date",
    ],
    "vendors": [
        "id", "name", "contact_email", "currency", "promised_lead_days",
        "actual_lead_days", "orders_total", "orders_complete", "damaged_units",
        "quality_score", "active",
    ],
    "vendor_prices": ["id", "vendor_id", "software_id", "price", "currency", "effective_date"],
    "license_pools": [
        "id", "software_id", "batch_code", "quantity", "assigned",
        "start_date", "expiry_date", "allocation_method", "status",
    ],
    "allocations": [
        "id", "pool_id", "software_id", "customer_id", "site", "tenant",
        "environment", "quantity", "allocated_at",
    ],
    "purchase_orders": [
        "id", "software_id", "vendor_id", "quantity", "unit_price", "currency",
        "total", "status", "requested_by", "approved_by", "created_at", "expected_date",
    ],
    "channels": [
        "id", "name", "type", "status", "last_sync", "products_synced",
        "licenses_synced", "sync_mode",
    ],
    "subscription_plans": [
        "id", "name", "monthly_price", "currency", "users_included",
        "locations_included", "features", "integration_addon_price", "active",
    ],
}


def _software_rows() -> list[list]:
    today = date.today()
    return [
        ["SW-001", "SecureFlow IAM", "CloudAxis", "Identity & Access", "Identity governance, SSO and privileged access for regulated enterprises.", "sso|mfa|identity governance|privileged access|role based access|audit trails", "banking|insurance|healthcare|government", "cloud|hybrid", "iso 27001|soc 2|gdpr|rbi", "per user", "USD", 74, 18, 1200, 890, today + timedelta(days=48), "Active", "USD", 1840000, 612000, 428000, 18, 7, "SF-PROD-1001", today.isoformat()],
        ["SW-002", "DataVista BI", "Northstar Analytics", "Analytics", "Self-service analytics and executive dashboards with broad data connectors.", "dashboards|data visualization|forecasting|excel integration|sql connectors|scheduled reports", "retail|manufacturing|banking|technology", "cloud|on-premise", "iso 27001|soc 2", "per user", "USD", 42, 16, 650, 412, today + timedelta(days=122), "Active", "USD", 985000, 286000, 194000, 14, 5, "SF-PROD-1002", (today - timedelta(days=1)).isoformat()],
        ["SW-003", "ServicePilot Pro", "OrbitWorks", "IT Service Management", "ITSM suite for incidents, assets, service catalog and workflow automation.", "incident management|asset management|service catalog|sla tracking|workflow automation|knowledge base", "technology|manufacturing|education|healthcare", "cloud|hybrid", "iso 27001|soc 2|hipaa", "per agent", "INR", 58000, 20, 420, 356, today + timedelta(days=29), "Active", "INR", 146500000, 48200000, 31600000, 22, 9, "SF-PROD-1003", today.isoformat()],
        ["SW-004", "RetailPulse CRM", "Mercury Labs", "CRM", "Omnichannel CRM with lead management, campaigns and customer service.", "lead management|campaigns|customer 360|whatsapp integration|email automation|case management", "retail|hospitality|real estate|ecommerce", "cloud", "gdpr|iso 27001", "per user", "INR", 36000, 15, 800, 486, today + timedelta(days=205), "Active", "INR", 89400000, 27300000, 22800000, 19, 8, "SF-PROD-1004", (today - timedelta(days=2)).isoformat()],
        ["SW-005", "VaultBackup Enterprise", "IronPeak", "Backup & Recovery", "Immutable backup and disaster recovery across cloud and data centers.", "immutable backup|disaster recovery|ransomware protection|vmware backup|database backup|retention policies", "banking|healthcare|government|manufacturing", "on-premise|hybrid|cloud", "iso 27001|soc 2|hipaa|rbi", "capacity based", "USD", 18500, 22, 90, 72, today + timedelta(days=14), "Active", "USD", 2210000, 748000, 615000, 16, 6, "SF-PROD-1005", today.isoformat()],
        ["SW-006", "PeopleCore HCM", "BlueOrbit", "Human Resources", "Unified HR, payroll, attendance and performance management platform.", "payroll|attendance|performance management|employee self service|recruitment|compliance reports", "manufacturing|retail|technology|services", "cloud|on-premise", "iso 27001|gdpr", "per employee", "INR", 7200, 14, 5000, 3710, today + timedelta(days=310), "Active", "INR", 118600000, 39200000, 26700000, 27, 11, "SF-PROD-1006", (today - timedelta(days=1)).isoformat()],
        ["SW-007", "FactoryIQ MES", "Vertex Industrial", "Manufacturing", "Manufacturing execution and OEE intelligence for connected factories.", "production planning|oee|quality management|iot integration|downtime tracking|traceability", "manufacturing|automotive|pharmaceutical", "on-premise|hybrid", "iso 27001|gmp", "per site", "USD", 48000, 21, 35, 27, today + timedelta(days=76), "Active", "USD", 3620000, 1240000, 960000, 11, 5, "SF-PROD-1007", today.isoformat()],
        ["SW-008", "ContractSphere CLM", "LexiCore", "Legal Operations", "Contract lifecycle management with approvals and obligation tracking.", "contract authoring|digital signature|approval workflow|obligation tracking|clause library|renewal alerts", "banking|technology|services|healthcare", "cloud", "iso 27001|soc 2|gdpr", "per user", "USD", 56, 17, 300, 118, today + timedelta(days=164), "Active", "USD", 640000, 168000, 212000, 9, 6, "SF-PROD-1008", (today - timedelta(days=3)).isoformat()],
        ["SW-009", "CloudGuard CSPM", "Aegis Cloud", "Cloud Security", "Multi-cloud posture management and continuous compliance monitoring.", "cloud posture management|misconfiguration detection|aws|azure|gcp|continuous compliance|risk scoring", "banking|technology|retail|government", "cloud", "iso 27001|soc 2|pci dss|rbi", "per cloud account", "USD", 7200, 19, 140, 61, today + timedelta(days=39), "Active", "USD", 1560000, 526000, 704000, 13, 10, "SF-PROD-1009", today.isoformat()],
        ["SW-010", "DocuMesh DMS", "Paperless Systems", "Document Management", "Secure document management with OCR, records retention and approvals.", "document management|ocr|version control|approval workflow|records retention|full text search", "government|legal|healthcare|education", "cloud|on-premise", "iso 27001|hipaa|gdpr", "per user", "INR", 18500, 15, 900, 529, today + timedelta(days=248), "Review", "INR", 67400000, 19400000, 15800000, 15, 7, "SF-PROD-1010", (today - timedelta(days=2)).isoformat()],
    ]


def _customer_rows() -> list[list]:
    return [
        ["CUS-001", "Apex Cooperative Bank", "banking", "India", "INR", "Priya Menon", "priya@apexbank.demo"],
        ["CUS-002", "Helio Manufacturing", "manufacturing", "India", "INR", "Arjun Shah", "arjun@helio.demo"],
        ["CUS-003", "Northwind Retail Group", "retail", "United States", "USD", "Maya Brooks", "maya@northwind.demo"],
        ["CUS-004", "Medora Health Network", "healthcare", "United States", "USD", "Daniel Kim", "daniel@medora.demo"],
        ["CUS-005", "CivicWorks Department", "government", "India", "INR", "Neha Rao", "neha@civicworks.demo"],
    ]


def _query_rows() -> list[list]:
    now = date.today()
    return [
        ["QRY-1001", "CUS-001", "Identity governance for 900 banking users", "We need SSO, MFA, privileged access controls and detailed audit trails for about 900 users. The solution should satisfy RBI and ISO 27001 expectations and support hybrid deployment.", now.isoformat(), "High", "New", 7200000, "INR", "hybrid", "sso|mfa|privileged access|audit trails|identity governance"],
        ["QRY-1002", "CUS-002", "Factory production and downtime visibility", "Our three plants need real-time OEE dashboards, downtime reasons, production planning, traceability and IoT machine integration. On-premise is preferred.", (now - timedelta(days=1)).isoformat(), "High", "New", 175000, "USD", "on-premise", "oee|downtime tracking|production planning|traceability|iot integration"],
        ["QRY-1003", "CUS-003", "Customer engagement platform", "Looking for a cloud CRM for 450 users with customer 360, lead management, email campaigns, WhatsApp integration and service case management.", (now - timedelta(days=2)).isoformat(), "Medium", "New", 310000, "USD", "cloud", "customer 360|lead management|email automation|whatsapp integration|case management"],
        ["QRY-1004", "CUS-004", "Ransomware-resilient clinical backup", "We need immutable backup, ransomware recovery, database and VMware protection, HIPAA controls and a hybrid deployment model.", (now - timedelta(days=3)).isoformat(), "Critical", "New", 120000, "USD", "hybrid", "immutable backup|ransomware protection|database backup|vmware backup|disaster recovery"],
        ["QRY-1005", "CUS-005", "Digitize records and approval workflows", "Please recommend an on-premise document platform with OCR, version control, retention policies, full-text search and departmental approval workflows.", (now - timedelta(days=5)).isoformat(), "Medium", "New", 4800000, "INR", "on-premise", "document management|ocr|version control|records retention|approval workflow|full text search"],
        ["QRY-1006", "CUS-001", "Continuous cloud compliance for AWS and Azure", "The bank requires continuous cloud compliance, risk scoring and misconfiguration detection across AWS and Azure with RBI and PCI DSS coverage.", (now - timedelta(days=7)).isoformat(), "High", "Analyzed", 85000, "USD", "cloud", "cloud posture management|misconfiguration detection|aws|azure|continuous compliance|risk scoring"],
    ]


def create_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    content = {
        "software": _software_rows(),
        "customers": _customer_rows(),
        "queries": _query_rows(),
        "users": [
            ["USR-001", "Aarav Admin", "admin@lean2automate.demo", "demo123", "admin", "", True],
            ["USR-002", "Sara Sales", "sales@lean2automate.demo", "demo123", "sales", "", True],
            ["USR-003", "Vikram Viewer", "viewer@lean2automate.demo", "demo123", "viewer", "", True],
            ["USR-004", "Priya Menon", "priya@apexbank.demo", "demo123", "customer", "CUS-001", True],
        ],
        "analyses": [],
        "opportunities": [
            ["OPP-001", "QRY-1006", "CUS-001", "SW-009", "Apex Bank cloud compliance", "Sara Sales", "Proposal", 65, 6800000, "INR", (date.today() + timedelta(days=32)).isoformat(), "Portal", "SF-OPP-2001", date.today().isoformat()],
            ["OPP-002", "QRY-1004", "CUS-004", "SW-005", "Medora ransomware recovery", "Sara Sales", "Qualification", 35, 120000, "USD", (date.today() + timedelta(days=58)).isoformat(), "Portal", "", date.today().isoformat()],
            ["OPP-003", "QRY-1002", "CUS-002", "SW-007", "Helio connected factories", "Aarav Admin", "Negotiation", 80, 175000, "USD", (date.today() + timedelta(days=21)).isoformat(), "Portal", "SF-OPP-2003", date.today().isoformat()],
        ],
        "feedback": [],
        "alerts": [
            ["ALT-001", "renewal", "critical", "VaultBackup renewal due", "Vendor agreement renews within 30 days.", "SW-005", (date.today() + timedelta(days=14)).isoformat(), "Open", date.today().isoformat()],
            ["ALT-002", "utilization", "warning", "ContractSphere underutilized", "License utilization is below 40%.", "SW-008", "", "Open", date.today().isoformat()],
        ],
        "audit_logs": [],
        "settings": [
            ["salesforce_mode", "demo", "Salesforce connector mode: demo or live"],
            ["email_provider", "simulated", "Email source: simulated, gmail, or outlook"],
            ["renewal_alert_days", "90", "Days before renewal to create an alert"],
            ["default_currency", "INR", "Portfolio reporting currency"],
        ],
        "email_inbox": [
            ["MAIL-001", "simulated", "MSG-9001", "procurement@zenithpharma.demo", "GMP manufacturing traceability platform", "We need production traceability, quality management, OEE and on-premise deployment across two pharmaceutical facilities. Budget USD 190000.", date.today().isoformat(), "Unread", ""],
            ["MAIL-002", "simulated", "MSG-9002", "it@brightlearn.demo", "Document management for education group", "Please suggest a cloud document platform with OCR, version control, full text search and approval workflows. Budget INR 3500000.", (date.today() - timedelta(days=1)).isoformat(), "Unread", ""],
        ],
        "procurement": [
            ["SW-001", 420, 8500, 3100, 18, 60, 120, (date.today() - timedelta(days=12)).isoformat()],
            ["SW-002", 210, 6200, 1700, 24, 35, 70, (date.today() - timedelta(days=98)).isoformat()],
            ["SW-003", 180, 12000, 9000, 21, 30, 65, (date.today() - timedelta(days=19)).isoformat()],
            ["SW-004", 260, 7600, 5400, 16, 45, 90, (date.today() - timedelta(days=41)).isoformat()],
            ["SW-005", 38, 18000, 42000, 35, 8, 15, (date.today() - timedelta(days=8)).isoformat()],
            ["SW-006", 850, 5000, 1050, 14, 120, 220, (date.today() - timedelta(days=27)).isoformat()],
            ["SW-007", 14, 24000, 92000, 48, 4, 8, (date.today() - timedelta(days=16)).isoformat()],
            ["SW-008", 90, 5800, 2200, 28, 18, 35, (date.today() - timedelta(days=124)).isoformat()],
            ["SW-009", 52, 15000, 18000, 30, 12, 24, (date.today() - timedelta(days=11)).isoformat()],
            ["SW-010", 230, 6800, 2800, 20, 40, 80, (date.today() - timedelta(days=94)).isoformat()],
        ],
        "vendors": [
            ["VEN-001", "CloudAxis", "orders@cloudaxis.demo", "USD", 18, 17, 24, 23, 2, 94, True],
            ["VEN-002", "Northstar Analytics", "supply@northstar.demo", "USD", 24, 29, 18, 15, 4, 78, True],
            ["VEN-003", "OrbitWorks", "commercial@orbitworks.demo", "INR", 21, 20, 31, 30, 1, 96, True],
            ["VEN-004", "Mercury Labs", "partners@mercury.demo", "INR", 16, 18, 20, 18, 3, 86, True],
            ["VEN-005", "IronPeak", "renewals@ironpeak.demo", "USD", 35, 32, 17, 17, 0, 98, True],
            ["VEN-006", "BlueOrbit", "sales@blueorbit.demo", "INR", 14, 15, 28, 27, 2, 92, True],
            ["VEN-007", "Vertex Industrial", "supply@vertex.demo", "USD", 48, 55, 12, 10, 2, 75, True],
            ["VEN-008", "LexiCore", "channel@lexicore.demo", "USD", 28, 34, 14, 11, 3, 72, True],
            ["VEN-009", "Aegis Cloud", "partners@aegis.demo", "USD", 30, 27, 19, 19, 1, 97, True],
            ["VEN-010", "Paperless Systems", "orders@paperless.demo", "INR", 20, 22, 22, 20, 2, 88, True],
        ],
        "vendor_prices": [
            ["VPR-001", "VEN-001", "SW-001", 70, "USD", "2025-09-01"], ["VPR-002", "VEN-001", "SW-001", 72, "USD", "2026-01-01"], ["VPR-003", "VEN-001", "SW-001", 74, "USD", "2026-05-01"],
            ["VPR-004", "VEN-005", "SW-005", 17200, "USD", "2025-08-01"], ["VPR-005", "VEN-005", "SW-005", 17800, "USD", "2026-01-01"], ["VPR-006", "VEN-005", "SW-005", 18500, "USD", "2026-05-01"],
            ["VPR-007", "VEN-007", "SW-007", 45500, "USD", "2025-10-01"], ["VPR-008", "VEN-007", "SW-007", 47000, "USD", "2026-02-01"], ["VPR-009", "VEN-007", "SW-007", 48000, "USD", "2026-05-01"],
        ],
        "license_pools": [
            ["POOL-001", "SW-001", "IAM-2026-A", 600, 510, "2026-01-01", "2026-12-31", "FIFO", "Active"],
            ["POOL-002", "SW-001", "IAM-2026-B", 600, 380, "2026-04-01", "2027-03-31", "FIFO", "Active"],
            ["POOL-003", "SW-005", "VBK-2025-R", 40, 35, "2025-07-01", (date.today() + timedelta(days=32)).isoformat(), "FIFO", "Expiring"],
            ["POOL-004", "SW-005", "VBK-2026-A", 50, 37, "2026-03-01", "2027-02-28", "FIFO", "Active"],
            ["POOL-005", "SW-008", "CLM-2025-A", 150, 48, "2025-07-01", (date.today() + timedelta(days=46)).isoformat(), "FIFO", "Expiring"],
            ["POOL-006", "SW-008", "CLM-2026-A", 150, 70, "2026-02-01", "2027-01-31", "FIFO", "Active"],
        ],
        "allocations": [
            ["ALC-001", "POOL-001", "SW-001", "CUS-001", "Mumbai HQ", "apex-prod", "Production", 420, "2026-01-15"],
            ["ALC-002", "POOL-001", "SW-001", "CUS-004", "New York Clinical", "medora-iam", "Production", 90, "2026-02-10"],
            ["ALC-003", "POOL-003", "SW-005", "CUS-004", "Chicago DR Site", "medora-backup", "Production", 20, "2025-08-12"],
            ["ALC-004", "POOL-003", "SW-005", "CUS-001", "Pune DR Center", "apex-dr", "Disaster Recovery", 15, "2025-09-05"],
            ["ALC-005", "POOL-005", "SW-008", "CUS-001", "Mumbai Legal", "apex-legal", "Production", 32, "2025-08-20"],
            ["ALC-006", "POOL-005", "SW-008", "CUS-003", "Seattle HQ", "northwind-legal", "Production", 16, "2025-10-11"],
        ],
        "purchase_orders": [],
        "channels": [
            ["CHN-001", "Salesforce CRM", "CRM", "Connected", date.today().isoformat(), 10, 4260, "Scheduled"],
            ["CHN-002", "Customer Self-Service", "Portal", "Connected", date.today().isoformat(), 10, 1850, "Real-time"],
            ["CHN-003", "Partner Marketplace", "Marketplace", "Attention", (date.today() - timedelta(days=2)).isoformat(), 7, 920, "Scheduled"],
            ["CHN-004", "Finance ERP", "ERP", "Connected", date.today().isoformat(), 10, 4260, "Nightly"],
        ],
        "subscription_plans": [
            ["PLAN-001", "Starter", 24999, "INR", 3, 1, "inventory|queries|basic matching|email support", 7500, True],
            ["PLAN-002", "Growth", 74999, "INR", 15, 5, "all starter|forecasting|comparison|alerts|proposal pdf", 15000, True],
            ["PLAN-003", "Enterprise", 199999, "INR", 100, 25, "all growth|salesforce|gmail|outlook|audit|custom roles", 35000, True],
        ],
    }

    for sheet_name, headers in HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in content[sheet_name]:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="172554")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = width

    workbook.save(path)


def migrate_workbook(path: Path) -> None:
    """Add new prototype sheets/columns without replacing existing user data."""
    if not path.exists():
        create_workbook(path)
        return
    workbook = load_workbook(path)
    seed_path = path.with_name(f".{path.stem}-schema-{uuid4().hex}.xlsx")
    try:
        create_workbook(seed_path)
        seed = load_workbook(seed_path, data_only=True)
        changed = False
        for sheet_name, headers in HEADERS.items():
            if sheet_name not in workbook.sheetnames:
                source = seed[sheet_name]
                target = workbook.create_sheet(sheet_name)
                for row in source.iter_rows(values_only=True):
                    target.append(list(row))
                changed = True
                continue
            sheet = workbook[sheet_name]
            existing_headers = [cell.value for cell in sheet[1]]
            source = seed[sheet_name]
            for header in headers:
                if header not in existing_headers:
                    sheet.cell(1, sheet.max_column + 1).value = header
                    existing_headers.append(header)
                    source_headers = [cell.value for cell in source[1]]
                    if header in source_headers:
                        source_column = source_headers.index(header) + 1
                        for row_index in range(2, min(sheet.max_row, source.max_row) + 1):
                            sheet.cell(row_index, len(existing_headers)).value = source.cell(row_index, source_column).value
                    changed = True
        if changed:
            workbook.save(path)
        seed.close()
    finally:
        workbook.close()
        seed_path.unlink(missing_ok=True)
