from __future__ import annotations

from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import load_workbook

from .seed_data import HEADERS, create_workbook, migrate_workbook


class ExcelRepository:
    def __init__(self, path: Path):
        self.path = path
        self._lock = Lock()
        if not self.path.exists():
            create_workbook(self.path)
        else:
            try:
                migrate_workbook(self.path)
            except PermissionError:
                # Excel can lock the workbook on Windows. Migration retries on restart.
                pass

    def all(self, sheet_name: str) -> list[dict[str, Any]]:
        with self._lock:
            workbook = load_workbook(self.path, data_only=True)
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
        if not rows:
            return []
        headers = [str(value) for value in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]

    def find(self, sheet_name: str, key: str, value: Any) -> dict[str, Any] | None:
        return next((row for row in self.all(sheet_name) if str(row.get(key)) == str(value)), None)

    def append(self, sheet_name: str, item: dict[str, Any]) -> None:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            sheet.append([item.get(header, "") for header in headers])
            workbook.save(self.path)
            workbook.close()

    def delete(self, sheet_name: str, key: str, value: Any) -> bool:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook[sheet_name]
            headers = {cell.value: cell.column for cell in sheet[1]}
            for row_index in range(2, sheet.max_row + 1):
                if str(sheet.cell(row_index, headers[key]).value) == str(value):
                    sheet.delete_rows(row_index)
                    workbook.save(self.path)
                    workbook.close()
                    return True
            workbook.close()
            return False

    def next_id(self, sheet_name: str, prefix: str, key: str = "id") -> str:
        values = []
        for row in self.all(sheet_name):
            raw = str(row.get(key) or "")
            if raw.startswith(f"{prefix}-"):
                try:
                    values.append(int(raw.split("-")[-1]))
                except ValueError:
                    continue
        return f"{prefix}-{max(values, default=0) + 1:03d}"

    def update(self, sheet_name: str, key: str, value: Any, changes: dict[str, Any]) -> None:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook[sheet_name]
            headers = {cell.value: cell.column for cell in sheet[1]}
            for row_index in range(2, sheet.max_row + 1):
                if str(sheet.cell(row_index, headers[key]).value) == str(value):
                    for field, field_value in changes.items():
                        if field in headers:
                            sheet.cell(row_index, headers[field]).value = field_value
                    workbook.save(self.path)
                    workbook.close()
                    return
            workbook.close()
            raise KeyError(f"{sheet_name} row not found: {key}={value}")

    def replace_analysis(self, query_id: str, rows: list[dict[str, Any]]) -> None:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook["analyses"]
            headers = HEADERS["analyses"]
            existing = [
                list(row)
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] != query_id
            ]
            sheet.delete_rows(2, max(sheet.max_row - 1, 0))
            for row in existing:
                sheet.append(row)
            for item in rows:
                sheet.append([item.get(header, "") for header in headers])
            workbook.save(self.path)
            workbook.close()
