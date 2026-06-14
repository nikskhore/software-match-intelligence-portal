from __future__ import annotations

import io
import json
import os
import math
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .repository import ExcelRepository


USD_TO_INR = 84


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def audit(repository: ExcelRepository, user_id: str, action: str, entity_type: str, entity_id: str, details: Any = "") -> None:
    repository.append("audit_logs", {
        "id": repository.next_id("audit_logs", "AUD"),
        "user_id": user_id,
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "details": json.dumps(details, default=str) if not isinstance(details, str) else details,
        "created_at": now(),
    })


def export_software(repository: ExcelRepository) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "software"
    rows = repository.all("software")
    headers = list(rows[0].keys()) if rows else []
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_software(repository: ExcelRepository, content: bytes) -> int:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0
    headers = [str(value) for value in rows[0]]
    count = 0
    for values in rows[1:]:
        item = dict(zip(headers, values))
        if not item.get("id"):
            continue
        existing = repository.find("software", "id", item["id"])
        if existing:
            repository.update("software", "id", item["id"], item)
        else:
            repository.append("software", item)
        count += 1
    return count


def generate_proposal(repository: ExcelRepository, query_id: str) -> bytes:
    query = repository.find("queries", "id", query_id)
    customer = repository.find("customers", "id", query["customer_id"]) if query else None
    analyses = [row for row in repository.all("analyses") if row["query_id"] == query_id]
    if not query or not customer or not analyses:
        raise ValueError("Analyze the inquiry before generating a proposal")

    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("lean2automate", styles["Title"]),
        Paragraph("Software Recommendation Proposal", styles["Heading1"]),
        Spacer(1, 5 * mm),
        Paragraph(f"<b>Customer:</b> {customer['name']}", styles["BodyText"]),
        Paragraph(f"<b>Requirement:</b> {query['subject']}", styles["BodyText"]),
        Paragraph(f"<b>Budget:</b> {query['currency']} {float(query['budget']):,.0f}", styles["BodyText"]),
        Spacer(1, 5 * mm),
        Paragraph(str(analyses[0]["summary"]), styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]
    table_data = [["Rank", "Recommended software", "Match", "Estimated annual cost"]]
    for index, row in enumerate(sorted(analyses, key=lambda item: item["score"], reverse=True), 1):
        table_data.append([str(index), row["software_name"], f"{row['score']}%", f"{row['currency']} {float(row['annual_cost']):,.0f}"])
    table = Table(table_data, colWidths=[18 * mm, 68 * mm, 24 * mm, 50 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07111b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5df")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f7f8")]),
        ("PADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([table, Spacer(1, 7 * mm), Paragraph("Prepared by lean2automate Software Match Intelligence Portal.", styles["Italic"])])
    document.build(story)
    return buffer.getvalue()


def salesforce_sync(repository: ExcelRepository) -> dict[str, Any]:
    live = bool(os.getenv("SALESFORCE_INSTANCE_URL") and os.getenv("SALESFORCE_ACCESS_TOKEN"))
    synced_at = now()
    for product in repository.all("software"):
        repository.update("software", "id", product["id"], {"crm_last_sync": synced_at})
    for opportunity in repository.all("opportunities"):
        if not opportunity.get("crm_id"):
            repository.update("opportunities", "id", opportunity["id"], {"crm_id": f"SF-{opportunity['id']}", "updated_at": synced_at})
    return {
        "mode": "live-ready" if live else "demo",
        "products": len(repository.all("software")),
        "opportunities": len(repository.all("opportunities")),
        "synced_at": synced_at,
        "message": "Credentials detected; live API adapter is ready." if live else "Demo synchronization completed locally. Configure Salesforce credentials for live mode.",
    }


def ingest_simulated_email(repository: ExcelRepository, mail_id: str) -> dict[str, Any]:
    mail = repository.find("email_inbox", "id", mail_id)
    if not mail:
        raise KeyError(mail_id)
    if mail.get("query_id"):
        return repository.find("queries", "id", mail["query_id"])
    customer = repository.all("customers")[0]
    text = f"{mail['subject']} {mail['body']}".lower()
    currency = "USD" if "usd" in text else "INR"
    budget = 190000 if currency == "USD" else 3500000
    requirements = []
    vocabulary = ["oee", "traceability", "quality management", "ocr", "version control", "full text search", "approval workflow", "cloud", "on-premise"]
    for term in vocabulary:
        if term in text:
            requirements.append(term)
    query_id = repository.next_id("queries", "QRY")
    query = {
        "id": query_id,
        "customer_id": customer["id"],
        "subject": mail["subject"],
        "email_body": mail["body"],
        "received_at": mail["received_at"],
        "priority": "Medium",
        "status": "New",
        "budget": budget,
        "currency": currency,
        "preferred_deployment": "on-premise" if "on-premise" in text else "cloud",
        "requirements": "|".join(requirements),
    }
    repository.append("queries", query)
    repository.update("email_inbox", "id", mail_id, {"status": "Imported", "query_id": query_id})
    return query


def generate_renewal_alerts(repository: ExcelRepository, days: int = 90) -> int:
    existing = {row["entity_id"] for row in repository.all("alerts") if row["type"] == "renewal" and row["status"] == "Open"}
    created = 0
    today = date.today()
    for product in repository.all("software"):
        renewal = product.get("renewal_date")
        if hasattr(renewal, "date"):
            renewal = renewal.date()
        if isinstance(renewal, datetime):
            renewal = renewal.date()
        if isinstance(renewal, str):
            renewal = date.fromisoformat(renewal[:10])
        if isinstance(renewal, date) and 0 <= (renewal - today).days <= days and product["id"] not in existing:
            repository.append("alerts", {
                "id": repository.next_id("alerts", "ALT"),
                "type": "renewal",
                "severity": "critical" if (renewal - today).days <= 30 else "warning",
                "title": f"{product['name']} renewal due",
                "message": f"Agreement renews in {(renewal - today).days} days.",
                "entity_id": product["id"],
                "due_date": renewal.isoformat(),
                "status": "Open",
                "created_at": now(),
            })
            created += 1
    return created


def procurement_analysis(repository: ExcelRepository) -> list[dict[str, Any]]:
    software = {row["id"]: row for row in repository.all("software")}
    results = []
    today = date.today()
    for row in repository.all("procurement"):
        product = software[row["software_id"]]
        demand = float(row["annual_demand"] or 0)
        order_cost = float(row["order_cost"] or 0)
        holding = float(row["holding_cost_per_unit"] or 0)
        eoq = math.sqrt((2 * demand * order_cost) / holding) if holding > 0 else 0
        available = int(product["available_licenses"] or 0) - int(product["assigned_licenses"] or 0)
        last_sale = row["last_sale_date"]
        if hasattr(last_sale, "date"):
            last_sale = last_sale.date()
        if isinstance(last_sale, datetime):
            last_sale = last_sale.date()
        if isinstance(last_sale, str):
            last_sale = date.fromisoformat(last_sale[:10])
        days_since_sale = (today - last_sale).days if isinstance(last_sale, date) else 0
        dead_stock = days_since_sale >= 90
        holding_cost = available * holding
        reorder = available <= int(row["reorder_point"] or 0)
        results.append({
            **row, "software_name": product["name"], "vendor": product["vendor"],
            "currency": product["currency"], "unit_price": product["unit_license_cost"],
            "available_units": available, "eoq": round(eoq), "reorder_required": reorder,
            "recommended_order": round(max(eoq, int(row["reorder_point"] or 0) - available)) if reorder else 0,
            "days_since_sale": days_since_sale, "dead_stock": dead_stock,
            "annual_holding_cost": round(holding_cost),
        })
    return results


def vendor_scorecards(repository: ExcelRepository) -> list[dict[str, Any]]:
    price_rows = repository.all("vendor_prices")
    results = []
    for vendor in repository.all("vendors"):
        delivery_score = max(0, min(100, 100 - max(0, float(vendor["actual_lead_days"]) - float(vendor["promised_lead_days"])) * 5))
        fulfillment = 100 * float(vendor["orders_complete"] or 0) / max(float(vendor["orders_total"] or 1), 1)
        damage_score = max(0, 100 - float(vendor["damaged_units"] or 0) * 4)
        overall = round(delivery_score * .35 + fulfillment * .35 + damage_score * .15 + float(vendor["quality_score"] or 0) * .15)
        prices = sorted([row for row in price_rows if row["vendor_id"] == vendor["id"]], key=lambda row: str(row["effective_date"]))
        fluctuation = 0
        if len(prices) > 1 and float(prices[0]["price"]):
            fluctuation = round((float(prices[-1]["price"]) - float(prices[0]["price"])) / float(prices[0]["price"]) * 100, 1)
        results.append({
            **vendor, "delivery_score": round(delivery_score), "fulfillment_rate": round(fulfillment, 1),
            "damage_score": round(damage_score), "overall_score": overall,
            "price_fluctuation": fluctuation, "price_history": prices,
        })
    return sorted(results, key=lambda row: row["overall_score"], reverse=True)


def generate_po_pdf(repository: ExcelRepository, po: dict[str, Any]) -> bytes:
    product = repository.find("software", "id", po["software_id"])
    vendor = repository.find("vendors", "id", po["vendor_id"])
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    data = [
        ["Purchase Order", po["id"]],
        ["Supplier", vendor["name"]],
        ["Supplier email", vendor["contact_email"]],
        ["Software", product["name"]],
        ["Quantity", str(po["quantity"])],
        ["Unit price", f"{po['currency']} {float(po['unit_price']):,.2f}"],
        ["Total", f"{po['currency']} {float(po['total']):,.2f}"],
        ["Expected delivery", str(po["expected_date"])],
        ["Approval status", po["status"]],
    ]
    table = Table(data, colWidths=[55 * mm, 110 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07111b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), .5, colors.HexColor("#cbd5df")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("PADDING", (0, 0), (-1, -1), 9),
    ]))
    document.build([Paragraph("lean2automate", styles["Title"]), Paragraph("Software License Purchase Order", styles["Heading1"]), Spacer(1, 6 * mm), table])
    return buffer.getvalue()


def roi_simulation(repository: ExcelRepository, holding_reduction: float, stockout_reduction: float, utilization_gain: float) -> dict[str, Any]:
    analysis = procurement_analysis(repository)
    current_holding = sum(row["annual_holding_cost"] for row in analysis)
    pipeline = sum(float(row.get("pipeline_value") or 0) * (USD_TO_INR if row.get("revenue_currency") == "USD" else 1) for row in repository.all("software"))
    inventory_value = sum(float(row["unit_license_cost"] or 0) * max(int(row["available_licenses"] or 0) - int(row["assigned_licenses"] or 0), 0) * (USD_TO_INR if row["currency"] == "USD" else 1) for row in repository.all("software"))
    holding_savings = current_holding * holding_reduction / 100
    stockout_recovery = pipeline * .08 * stockout_reduction / 100
    utilization_recovery = inventory_value * utilization_gain / 100
    total = holding_savings + stockout_recovery + utilization_recovery
    return {
        "current_holding_cost_inr": round(current_holding),
        "holding_savings_inr": round(holding_savings),
        "stockout_recovery_inr": round(stockout_recovery),
        "utilization_recovery_inr": round(utilization_recovery),
        "total_reclaimed_cash_inr": round(total),
        "annual_roi_percent": round(total / max(inventory_value, 1) * 100, 1),
    }
